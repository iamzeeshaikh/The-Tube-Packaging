#!/usr/bin/env python3
"""Flatten the Search Console workbook into CSVs the baseline crawler can join.

The export is one workbook per property with a sheet per dimension. GSC exports
Queries and Pages as *separate* tables with no query-to-page pairing, which is
why nothing downstream can claim to have measured which page ranks for which
query.

Usage: python3 scripts/gsc-export.py "<path to .xlsx>"
"""
import csv
import pathlib
import sys

import openpyxl

OUT = pathlib.Path(__file__).resolve().parent.parent / "data" / "gsc"


def main(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    OUT.mkdir(parents=True, exist_ok=True)
    for name in wb.sheetnames:
        if name == "Chart":
            continue
        ws = wb[name]
        rows = [[c for c in r] for r in ws.iter_rows(values_only=True)]
        rows = [r for r in rows if any(c is not None for c in r)]
        if not rows:
            continue
        dest = OUT / (name.lower().replace(" ", "-") + ".csv")
        with dest.open("w", newline="", encoding="utf-8") as fh:
            csv.writer(fh).writerows(rows)
        print(f"{dest.name:24} {len(rows) - 1:>5} rows")


if __name__ == "__main__":
    main(sys.argv[1])
